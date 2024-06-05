import streamlit as st
import pandas as pd
import re
from PyPDF2 import PdfReader
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import base64
import requests
import json

def sanitize_sheet_name(name):
    return re.sub(r'[\\/*?:"<>|]', "", name)

def parse_pdf(file):
    # Initialize PDF reader
    reader = PdfReader(file)

    # Initialize list of dataframes for each sheet
    sheets = {}
    current_discipline = None
    current_data = []
    columns = ["Member Federation", "Surname", "First Name", "DoB", "PB", "SB"]

    # Iterate through each page
    for page_num in range(len(reader.pages)):
        page = reader.pages[page_num]
        text = page.extract_text()

        # Split text by lines
        lines = text.split('\n')

        # Parse each line
        for line in lines:
            if 'FINAL ENTRIES' in line:
                # Save the previous discipline sheet if it exists
                if current_discipline and current_data:
                    sheets[current_discipline] = current_data

                # Extract the discipline name
                discipline_match = re.search(r'FINAL ENTRIES\n(.*?)(?= Num\. of countries:)', text)
                if discipline_match:
                    current_discipline = sanitize_sheet_name(discipline_match.group(1).strip())
                    current_data = []

            elif current_discipline:
                # Add athlete information to the current discipline
                parts = line.split()

                # Check for valid row with date of birth (DoB) field
                if len(parts) >= 6 and re.match(r'\d{2}/\d{2}/\d{4}', parts[-4]):
                    # Combine Member Federation if it contains (R)
                    if parts[1] == '(R)':
                        member_federation = f"{parts[0]} {parts[1]}"
                        parts = [member_federation] + parts[2:]
                    else:
                        member_federation = parts[0]

                    surname = parts[1]
                    first_name = ' '.join(parts[2:-4])
                    dob = parts[-4]

                    # Determine PB and SB values, including any suffix
                    pb = parts[-3]
                    pb_suffix = parts[-2] if parts[-2] in ["sh", "i"] else ""
                    sb = parts[-1] if parts[-2] in ["sh", "i"] else parts[-2]
                    sb_suffix = parts[-1] if parts[-1] in ["sh", "i"] else ""

                    if pb_suffix:
                        pb += f" {pb_suffix}"
                    if sb_suffix:
                        sb += f" {sb_suffix}"

                    current_data.append([member_federation, surname, first_name, dob, pb, sb])
                elif len(parts) == 6 and re.match(r'\d{2}/\d{2}/\d{4}', parts[3]):
                    member_federation = parts[0]
                    surname = parts[1]
                    first_name = parts[2]
                    dob = parts[3]
                    pb = parts[4]
                    sb = parts[5]
                    current_data.append([member_federation, surname, first_name, dob, pb, sb])
                elif len(parts) == 5 and re.match(r'\d{2}/\d{2}/\d{4}', parts[3]):
                    member_federation = parts[0]
                    surname = parts[1]
                    first_name = parts[2]
                    dob = parts[3]
                    pb = parts[4]
                    sb = ''
                    current_data.append([member_federation, surname, first_name, dob, pb, sb])
                elif len(parts) == 4 and re.match(r'\d{2}/\d{2}/\d{4}', parts[2]):
                    member_federation = parts[0]
                    surname = parts[1]
                    first_name = ''
                    dob = parts[2]
                    pb = parts[3]
                    sb = ''
                    current_data.append([member_federation, surname, first_name, dob, pb, sb])
                elif len(parts) >= 3 and re.match(r'\d{2}/\d{2}/\d{4}', parts[-3]):
                    member_federation = parts[0]
                    surname = parts[1]
                    first_name = ' '.join(parts[2:-3])
                    dob = parts[-3]
                    pb = parts[-2]
                    sb = parts[-1]
                    current_data.append([member_federation, surname, first_name, dob, pb, sb])
                else:
                    # Handle cases with PB or SB values containing additional characters like "sh" or "i"
                    dob_index = next((i for i, part in enumerate(parts) if re.match(r'\d{2}/\d{2}/\d{4}', part)), None)
                    if dob_index:
                        member_federation = parts[0]
                        surname = parts[1]
                        first_name = ' '.join(parts[2:dob_index])
                        dob = parts[dob_index]
                        pb = ' '.join(parts[dob_index + 1:dob_index + 2]) if len(parts) > dob_index + 1 else ''
                        sb = ' '.join(parts[dob_index + 2:dob_index + 4]) if len(parts) > dob_index + 2 else ''
                        current_data.append([member_federation, surname, first_name, dob, pb, sb])

    # Add the last sheet if it exists
    if current_discipline and current_data:
        sheets[current_discipline] = current_data

    return sheets

def save_to_excel(sheets, output_path):
    # Create a new Excel file
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for discipline, data in sheets.items():
            df = pd.DataFrame(data, columns=["Member Federation", "Surname", "First Name", "DoB", "PB", "SB"])
            df.to_excel(writer, sheet_name=discipline, index=False)

    # Load the created Excel file
    wb = load_workbook(output_path)

    for discipline in sheets.keys():
        ws = wb[discipline]
        row_count = ws.max_row

        # Write the count at the bottom
        ws.cell(row=row_count + 2, column=1, value='Num. Athletes')
        ws.cell(row=row_count + 2, column=2, value=row_count - 1)  # Subtract 1 to exclude header

        # Optional: adjust cell alignment for better readability
        ws.cell(row=row_count + 2, column=1).alignment = Alignment(horizontal="right")
        ws.cell(row=row_count + 2, column=2).alignment = Alignment(horizontal="center")

    # Save the workbook with the new content
    wb.save(output_path)

    return output_path

def push_to_github(file_path, repo, branch, token, commit_message):
    # Read the file content
    with open(file_path, "rb") as f:
        content = f.read()

    # Encode the file content to base64
    content_base64 = base64.b64encode(content).decode()

    # Create the API URL
    url = f"https://api.github.com/repos/{repo}/contents/{file_path}"

    # Create the payload
    payload = {
        "message": commit_message,
        "branch": branch,
        "content": content_base64
    }

    # Set the headers
    headers = {
        "Authorization": f"token {token}",
        "Accept": "application/vnd.github.v3+json"
    }

    # Make the API request
    response = requests.put(url, headers=headers, data=json.dumps(payload))

    # Check the response
    if response.status_code == 201:
        print("File successfully pushed to GitHub.")
    else:
        print(f"Failed to push file to GitHub: {response.json()}")

# Streamlit app
st.title("PDF to Excel Converter")

uploaded_file = st.file_uploader("Choose a PDF file", type="pdf")

if uploaded_file is not None:
    # Parse the PDF
    sheets = parse_pdf(uploaded_file)

    if sheets:
        # Save to Excel
        output_path = "Rome2024_EM_finalentries.xlsx"
        file_path = save_to_excel(sheets, output_path)

        # Provide download link
        st.success(f"Excel file created successfully: {output_path}")
        with open(output_path, "rb") as file:
            st.download_button(label="Download Excel file", data=file, file_name=output_path, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        # GitHub credentials and info
        repo = st.text_input("GitHub Repository (owner/repo)", value="owner/repo")
        branch = st
